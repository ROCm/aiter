import openpyxl
import csv
from dataclasses import dataclass, asdict
import argparse


@dataclass
class BaseRecord:
    M: int
    N: int
    K: int


@dataclass
class Record(BaseRecord):
    bias: bool = False
    dtype: str = "torch.bfloat16"
    outdtype: str = "torch.bfloat16"
    scaleAB: bool = False


def to_record(
    mnkRecord: BaseRecord,
    bias=False,
    dtype="torch.bfloat16",
    outdtype="torch.bfloat16",
    scaleAB=False,
) -> Record:
    return Record(
        M=mnkRecord.M,
        N=mnkRecord.N,
        K=mnkRecord.K,
        bias=bias,
        dtype=dtype,
        outdtype=outdtype,
        scaleAB=scaleAB,
    )


def save_gemm_benchmark_result(records, csv_file_name):
    import csv
    from dataclasses import asdict

    csv_file = f"{csv_file_name}.csv"
    fieldnames = [
        "M",
        "N",
        "K",
        "TP",
        "quant_method",
        "quant_type",
        "output_type",
        "latency",
        "throughput",
        "bandwidth",
        "latency_asm",
        "throughput_asm",
        "bandwidth_asm",
        "latency_triton",
        "throughput_triton",
        "bandwidth_triton",
    ]

    print("====records={}".format(records))
    with open(csv_file, mode="w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for rec in records:
            writer.writerow(asdict(rec))
    print(f"data write to {csv_file} success!!")


def excel_to_struct_list(excel_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' ???!")
    sheet = wb[sheet_name]

    baseRecords = []

    # read D,E,F column (M, N, K) from 6 row
    for row in sheet.iter_rows(min_row=6, min_col=4, max_col=6, values_only=True):
        if all(v is None for v in row):
            continue
        M, N, K = row
        mnkrecord = BaseRecord(M=M, N=N, K=K)
        baseRecords.append(mnkrecord)

    return baseRecords


def csv_to_struct_list(csv_file):
    import csv

    with open(csv_file, mode="r", encoding="utf-8-sig") as file:
        reader = csv.reader(file)
        baseRecords = []
        for idx, row in enumerate(reader):
            if idx > 0:
                M = int(row[0])
                N = int(row[1])
                K = int(row[2])
                mnkrecord = BaseRecord(M=M, N=N, K=K)
                baseRecords.append(mnkrecord)
    return baseRecords


def save_structs_to_csv(records, csv_file, fieldnames):
    with open(csv_file, mode="w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for rec in records:
            writer.writerow(asdict(rec))

    print(f"data write to {csv_file} success!!")


def create_argument_parser():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawTextHelpFormatter,
        description="convert the csv file to untuned_gemm format",
    )
    parser.add_argument(
        "-i",
        "--input_file",
        type=str,
        nargs="?",
        const=None,
        default=None,
        help="""The input file containing GEMM MNK shape.
        e.g.: -i Qwen3-32B.csv or Qwen3-32B.xlsx""",
    )
    parser.add_argument(
        "-s",
        "--sheep_name",
        type=str,
        nargs="?",
        const=None,
        default="Sheet1",
        help="""The sheet name in xlsx file.
        e.g.: -s Qwen32B-Gemm""",
    )
    parser.add_argument(
        "-o",
        "--output_name",
        type=str,
        nargs="?",
        const=None,
        default=None,
        help="""The output csv file matching the untuned GEMM csv file format.
        e.g.: -o qwen32B_untuned_gemm""",
    )
    return parser


def save_untuned_gemm_csv(input_file, output_name):
    mnkrecords = csv_to_struct_list(input_file)
    records_base = []
    records = []
    for mnk in mnkrecords:
        if mnk not in records_base:
            records_base.append(mnk)

        record = to_record(mnk)
        print("----record={}".format(record))
        print("----records={}".format(records))
        if record not in records:
            records.append(record)

    save_structs_to_csv(records_base, f"{output_name}.csv", ["M", "N", "K"])
    save_structs_to_csv(
        records,
        f"{output_name}_bf16.csv",
        ["M", "N", "K", "bias", "dtype", "outdtype", "scaleAB"],
    )


if __name__ == "__main__":
    # csv_file_name = "qwen32B_untuned_gemm"
    # csv_file="/home/hatwu/aiter/op_tests/module_tests/Qwen3-30B.csv"
    parser = create_argument_parser()
    args = parser.parse_args()
    if ".xlsx" in args.input_file:
        mnkrecords = excel_to_struct_list(args.input_file, args.sheet_name)
    elif ".csv" in args.input_file:
        mnkrecords = csv_to_struct_list(args.input_file)
    else:
        raise ValueError(
            f"We do not support the file convertion of '{args.input_file}'!"
        )

    records = [to_record(mnk) for mnk in mnkrecords]
    save_structs_to_csv(mnkrecords, f"{args.output_name}.csv", ["M", "N", "K"])
    save_structs_to_csv(
        records,
        f"{args.output_name}_bf16.csv",
        ["M", "N", "K", "bias", "dtype", "outdtype", "scaleAB"],
    )
